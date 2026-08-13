"""
Génération du rapport Excel client, au format du labo :
- Feuille DIAMETRES : diamètres/section, par gamme, côte à côte
- Une feuille par gamme (nommée d'après son code) : courbes complètes,
  un bloc de colonnes par échantillon
- Feuille RECAPITULATIF : tableau détaillé + tableau transposé de
  synthèse + graphiques en barres comparant les gammes

Entrée : une liste de ResultatGamme (le résultat déjà calculé, tel que
renvoyé par TractionCalculateur - pas besoin de recalculer quoi que ce
soit ici, uniquement mettre en forme ce qui existe déjà).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from moteur_python.calculs.outils.courbe import courbe_contrainte_deformation
from moteur_python.calculs.outils.statistiques import moyenne_et_ecart_type


def _nom_feuille_valide(nom: str) -> str:
    """Excel limite les noms de feuille à 31 caractères et interdit
    certains symboles - on nettoie pour éviter un crash à l'écriture."""
    interdits = "/\\?*[]:"
    for c in interdits:
        nom = nom.replace(c, "-")
    return nom[:31]


def _nom_gamme(resultat) -> str:
    """Nom lisible d'une gamme, utilisé comme en-tête et nom de feuille."""
    return resultat.gamme.materiau.code_interne or resultat.gamme.materiau.nom_usage


def _ecrire_feuille_diametres(wb: Workbook, resultats: list) -> None:
    ws = wb.create_sheet("DIAMETRES")
    col = 1
    for resultat in resultats:
        nom_gamme = _nom_gamme(resultat)
        ws.cell(row=1, column=col, value=nom_gamme).font = Font(bold=True)
        entetes = ["ID Éch.", "D1 (mm)", "D2 (mm)", "D3 (mm)", "D_moy (mm)", "S (mm²)"]
        for i, e in enumerate(entetes):
            ws.cell(row=2, column=col + i, value=e).font = Font(bold=True)

        r = 3
        for ech in resultat.gamme.echantillons:
            diametres = ech.diametres_mm
            d_moy = sum(diametres) / len(diametres) if diametres else None
            section = ech.section_mm2() if diametres or (ech.largeur_mm and ech.epaisseur_mm) else None
            ws.cell(row=r, column=col, value=ech.identifiant)
            for i in range(3):
                val = diametres[i] if i < len(diametres) else None
                ws.cell(row=r, column=col + 1 + i, value=val)
            ws.cell(row=r, column=col + 4, value=d_moy)
            ws.cell(row=r, column=col + 5, value=section)
            r += 1

        col += 7  # 6 colonnes de données + 1 colonne de séparation


def _ecrire_feuille_courbe_gamme(wb: Workbook, resultat) -> None:
    nom_gamme = _nom_gamme(resultat)
    ws = wb.create_sheet(_nom_feuille_valide(nom_gamme))

    resultats_par_id = {r.identifiant: r for r in resultat.resultats_echantillons}

    col = 1
    for ech in resultat.gamme.echantillons:
        entetes = ["Elong.(mm)", "Force(N)", "Def", "Cont", "Cont(max)", "Def rupt", "E(mpa)"]
        ws.cell(row=1, column=col, value=f"{entetes[0]} - {ech.identifiant}").font = Font(bold=True)
        for i, e in enumerate(entetes[1:], start=1):
            ws.cell(row=1, column=col + i, value=e).font = Font(bold=True)

        section = ech.section_mm2()
        points_transformes = courbe_contrainte_deformation(
            [(p.force_newton, p.deplacement_mm) for p in ech.points_courbe],
            section_mm2=section,
            longueur_mm=ech.longueur_initiale_mm,
        )

        r_res = resultats_par_id.get(ech.identifiant)
        for r, (pt, (deform, cont)) in enumerate(zip(ech.points_courbe, points_transformes), start=2):
            ws.cell(row=r, column=col, value=pt.deplacement_mm)
            ws.cell(row=r, column=col + 1, value=pt.force_newton)
            ws.cell(row=r, column=col + 2, value=deform)
            ws.cell(row=r, column=col + 3, value=cont)

        if r_res:
            ws.cell(row=2, column=col + 4, value=r_res.contrainte_rupture_mpa)
            ws.cell(row=2, column=col + 5, value=r_res.deformation_rupture_pourcent / 100)
            ws.cell(row=2, column=col + 6, value=r_res.module_young_mpa)

        col += 8  # 7 colonnes + 1 séparation


def _ecrire_feuille_recapitulatif(wb: Workbook, resultats: list) -> None:
    ws = wb.create_sheet("RECAPITULATIF")

    # --- Tableau détaillé par échantillon, un bloc de colonnes par gamme ---
    col_debuts = {}
    col = 1
    for resultat in resultats:
        nom_gamme = _nom_gamme(resultat)
        col_debuts[nom_gamme] = col
        ws.cell(row=1, column=col, value=nom_gamme).font = Font(bold=True)
        for i, e in enumerate(["Cont(max)", "Def rupt", "E(mpa)"], start=1):
            ws.cell(row=2, column=col + i, value=e).font = Font(bold=True)
        col += 5  # ID + 3 valeurs + séparation

    nb_max_ech = max(len(r.resultats_echantillons) for r in resultats)
    for i in range(nb_max_ech):
        r_ligne = 3 + i
        for resultat in resultats:
            nom_gamme = _nom_gamme(resultat)
            c = col_debuts[nom_gamme]
            if i < len(resultat.resultats_echantillons):
                res = resultat.resultats_echantillons[i]
                ws.cell(row=r_ligne, column=c, value=res.identifiant)
                ws.cell(row=r_ligne, column=c + 1, value=res.contrainte_rupture_mpa)
                ws.cell(row=r_ligne, column=c + 2, value=res.deformation_rupture_pourcent / 100)
                ws.cell(row=r_ligne, column=c + 3, value=res.module_young_mpa)

    ligne_moyenne = 3 + nb_max_ech
    ligne_ecart_type = ligne_moyenne + 1
    ws.cell(row=ligne_moyenne, column=1, value="MOYENNE").font = Font(bold=True)
    ws.cell(row=ligne_ecart_type, column=1, value="ECART-TYPE").font = Font(bold=True)

    stats_par_gamme = {}  # nom_gamme -> {"cont": (m,e), "def": (m,e), "mod": (m,e)}
    for resultat in resultats:
        nom_gamme = _nom_gamme(resultat)
        c = col_debuts[nom_gamme]
        conts = [r.contrainte_rupture_mpa for r in resultat.resultats_echantillons]
        defs = [r.deformation_rupture_pourcent / 100 for r in resultat.resultats_echantillons]
        mods = [r.module_young_mpa for r in resultat.resultats_echantillons]

        m_cont, e_cont = moyenne_et_ecart_type(conts)
        m_def, e_def = moyenne_et_ecart_type(defs)
        m_mod, e_mod = moyenne_et_ecart_type(mods)
        stats_par_gamme[nom_gamme] = {"cont": (m_cont, e_cont), "def": (m_def, e_def), "mod": (m_mod, e_mod)}

        ws.cell(row=ligne_moyenne, column=c, value=nom_gamme)
        ws.cell(row=ligne_moyenne, column=c + 1, value=m_cont)
        ws.cell(row=ligne_moyenne, column=c + 2, value=m_def)
        ws.cell(row=ligne_moyenne, column=c + 3, value=m_mod)
        ws.cell(row=ligne_ecart_type, column=c, value="ECART-TYPE")
        ws.cell(row=ligne_ecart_type, column=c + 1, value=e_cont)
        ws.cell(row=ligne_ecart_type, column=c + 2, value=e_def)
        ws.cell(row=ligne_ecart_type, column=c + 3, value=e_mod)

    # --- Tableau transposé de synthèse (source des graphiques) ---
    ligne_transpo_debut = ligne_ecart_type + 3
    noms_gammes = [_nom_gamme(r) for r in resultats]

    blocs_metriques = [
        ("Cont(max)", "cont", 1),
        ("Def rupt", "def", 1 + 4),
        ("E(mpa)", "mod", 1 + 8),
    ]
    for label, cle, col_bloc in blocs_metriques:
        ws.cell(row=ligne_transpo_debut, column=col_bloc, value=label).font = Font(bold=True)
        for i, nom_gamme in enumerate(noms_gammes):
            ws.cell(row=ligne_transpo_debut, column=col_bloc + 1 + i, value=nom_gamme).font = Font(bold=True)
        ws.cell(row=ligne_transpo_debut + 1, column=col_bloc, value=label)
        ws.cell(row=ligne_transpo_debut + 2, column=col_bloc, value="ECART-TYPE")
        for i, nom_gamme in enumerate(noms_gammes):
            moyenne, ecart = stats_par_gamme[nom_gamme][cle]
            ws.cell(row=ligne_transpo_debut + 1, column=col_bloc + 1 + i, value=moyenne)
            ws.cell(row=ligne_transpo_debut + 2, column=col_bloc + 1 + i, value=ecart)

    # --- Graphiques en barres (un par grandeur) ---
    ligne_graph = ligne_transpo_debut + 5
    for idx, (label, cle, col_bloc) in enumerate(blocs_metriques):
        chart = BarChart()
        chart.title = label
        chart.y_axis.title = label
        data = Reference(
            ws, min_col=col_bloc + 1, max_col=col_bloc + len(noms_gammes),
            min_row=ligne_transpo_debut + 1, max_row=ligne_transpo_debut + 1,
        )
        cats = Reference(
            ws, min_col=col_bloc + 1, max_col=col_bloc + len(noms_gammes),
            min_row=ligne_transpo_debut, max_row=ligne_transpo_debut,
        )
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 10
        ancre = f"{get_column_letter(1 + idx * 6)}{ligne_graph}"
        ws.add_chart(chart, ancre)


def generer_rapport_labo(resultats: list, chemin_sortie: str) -> None:
    """
    Point d'entrée principal.

    resultats : liste de ResultatGamme (au moins 1). Si un seul élément,
    le rapport ne contient qu'une seule colonne par tableau (pas de
    comparaison) - reste valide, juste moins riche.
    """
    if not resultats:
        raise ValueError("Aucun résultat fourni - impossible de générer un rapport.")

    wb = Workbook()
    wb.remove(wb.active)  # on retire la feuille par défaut, vide

    _ecrire_feuille_diametres(wb, resultats)
    for resultat in resultats:
        _ecrire_feuille_courbe_gamme(wb, resultat)
    _ecrire_feuille_recapitulatif(wb, resultats)

    wb.save(chemin_sortie)
